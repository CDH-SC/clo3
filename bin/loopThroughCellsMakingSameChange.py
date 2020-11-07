#! python3
# testScript.py - Loops through sheet, correcting one column; replacing it with substring

import openpyxl

# define constants
SHEET_NAME = 'Sheet1'
ROW_WITH_FIELD_TITLES = 1 
COL_CHANGING = 3 
EXCEL_FILE_NAME = 'Volume0.xlsx'
UPDATED_FILE_NAME = 'Volume0_updated.xlsx'
# openpyxl has an object of class workbook which contains the sheets within current working directory 
wb = openpyxl.load_workbook(EXCEL_FILE_NAME) 

# get the sheet 
sheet = wb[SHEET_NAME] 


# loop through sheet changing columns as required for colNum in range(COL

for rowNum in range(ROW_WITH_FIELD_TITLES+1, sheet.max_row+1): # skip first row 
	currVal = str(sheet.cell(row=rowNum,column=COL_CHANGING).value) 
	modifiedVal = currVal[7:]
	sheet.cell(row=rowNum,column=COL_CHANGING).value = modifiedVal

wb.save(UPDATED_FILE_NAME)	

for rowNum in range(ROW_WITH_FIELD_TITLES+1, sheet.max_row):
	print(str(sheet.cell(row=rowNum,column=COL_CHANGING).value))
