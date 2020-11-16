#! python3
# testScript.py - Loops through sheet, correcting one column; replacing it with substring

# Steps to running it:
	# 1. move it into directory of excel file currently working on
	# 2. change EXCEL_FILE_NAME constant to reflect excel file in cwd
	# 3. execute, then repeat steps 1-2
import openpyxl

# these constants are modified based on which column and rows looping through
SHEET_NAME = 'Sheet1'
ROW_WITH_FIELD_TITLES = 1 
COL_CHANGING = 3

# as of right now this, excel file name constant must be changed before moving onto the next Volume 
EXCEL_FILE_NAME = 'Volume0.xlsx'

# this script runs as expected so we can precautionary measure of adding changes to a different file initially
UPDATED_FILE_NAME = 'Volume0_updated.xlsx'

# openpyxl has an object of class workbook which contains the sheets within current working directory 
wb = openpyxl.load_workbook(EXCEL_FILE_NAME) 

# get the sheet 
sheet = wb[SHEET_NAME] 


# loop through sheet changing columns as required; starting at the second row
# for changing "Volume," substring to "Volume," will need to save two substrings of cell's value, combine it with a defined string of "Album,"
# if you make "Album," string a constant it'll run more efficiently but either way works and prob is only a split second difference

# Steps:
	# 1: save substring of characters {<first letter> TO <whitespace right before the 'V' in 'Volume'}
	# 2: save substring of characters {<first whitespace after 'e,' in 'Volume' TO end character}
	# 3: concatenate new string with firstSubStr + albumStr+ secondSubStr
	# 4: set the value of cell currently at in loop to the new string
for rowNum in range(ROW_WITH_FIELD_TITLES+1, sheet.max_row+1): # skip first row 
	currVal = str(sheet.cell(row=rowNum,column=COL_CHANGING).value) 
	modifiedVal = currVal[7:] # this is a substring from the seventh character to the end character
	sheet.cell(row=rowNum,column=COL_CHANGING).value = modifiedVal

wb.save(UPDATED_FILE_NAME)	

for rowNum in range(ROW_WITH_FIELD_TITLES+1, sheet.max_row):
	print(str(sheet.cell(row=rowNum,column=COL_CHANGING).value))
